import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


class ExcelWriter:
    def __init__(self, output_path):
        self.output_path = output_path
        self.writer = None

    def __enter__(self):
        self.writer = pd.ExcelWriter(self.output_path, engine='openpyxl')
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.writer:
            self.writer.close()

    def write_sheet(self, sheet_name, df):
        df.to_excel(self.writer, sheet_name=sheet_name, index=False)

    def write_summary_region(self, summary_df):
        summary_df.to_excel(self.writer, sheet_name='Summary_Region', index=False)

    def write_region_sheet(self, region_name, data_df):
        data_df.to_excel(self.writer, sheet_name=region_name, index=False)

    def write_open_po(self, overseas_df, china_df):
        overseas_df.to_excel(self.writer, sheet_name='Open PO', index=False)
        
        if not china_df.empty:
            workbook = self.writer.book
            worksheet = workbook['Open PO']
            
            start_col = len(overseas_df.columns) + 2
            
            for col_idx, col_name in enumerate(china_df.columns):
                worksheet.cell(row=1, column=start_col + col_idx, value=col_name)
            
            for row_idx, (_, row) in enumerate(china_df.iterrows(), start=2):
                for col_idx, value in enumerate(row):
                    worksheet.cell(row=row_idx, column=start_col + col_idx, value=value)

    def write_unmapped_report(self, unmapped_list):
        if not unmapped_list:
            return
        
        df = pd.DataFrame(unmapped_list)
        df.to_excel(self.writer, sheet_name='未映射报告', index=False)
